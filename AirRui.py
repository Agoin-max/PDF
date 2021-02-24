import pandas as pd
import openpyxl
import win32com.client as win32
import sys
import xlsxwriter
import time


class TranFormation:

    def tran_xlsx(self, fname):
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.DisplayAlerts = 0  # 不显示弹窗
        wb = excel.Workbooks.Open(fname)
        wb.SaveAs(fname + "x", FileFormat=51)
        wb.Close()
        excel.Workbooks.Close()
        excel.Application.Quit()


fname = r"C:\Users\windo\Desktop\JL82721057(2.1)HK.xls"
TranFormation().tran_xlsx(fname)


class NewModel:
    def __init__(self, uid=1, shopname="", brand="", model="", unit="个", place="",shopdescription = ""):
        self.uid = uid  # 序号
        self.shopname = shopname  # 货物中文名称
        self.brand = brand  # 品牌
        self.model = model  # 型号
        self.unit = unit  # 单位
        self.place = place  # 产地
        self.shopdescription = shopdescription

class AirModel:

    def __init__(self):
        self.__set_list = []
        self.__newmodel_list = []

    def invoke_openpyxl(self, path):
        wb = openpyxl.load_workbook(path)
        sheet = wb.sheetnames
        ws = wb[sheet[0]]
        ws.title = "ATM"
        max_row = ws.max_row
        max_cow = ws.max_column
        cow_brand = 0
        cow_PN = 0
        start_r = 0
        start_e = 0
        for r in range(1, max_row + 1):
            for c in range(1, max_cow + 1):
                if ws.cell(r, c).value == "序号":
                    start_r = r
                    start_e = c
                if ws.cell(r, c).value == "品牌":
                    cow_brand = c
                if ws.cell(r, c).value == "P/N":
                    cow_PN = c

        end_r = 0
        for r in range(start_r, max_row + 1):
            if ws.cell(r, start_e).value == None:
                end_r = r
                break

        cow_shopsmall = 0
        cow_shopname = 0
        cow_desc = 0
        for r in range(start_r, end_r):
            for c in range(1, max_cow + 1):
                if ws.cell(r, c).value == "Remark":
                    ws.cell(r, c + 1).value = "商品小类"
                    ws.cell(r, c + 2).value = "商品品名"
                    ws.cell(r, c + 3).value = "商品描述"
                    cow_shopsmall = c + 1
                    cow_shopname = c + 2
                    cow_desc = c + 3

        df = pd.read_excel(r"C:\Users\windo\Desktop\历史数据.xlsx")
        list_obj = []
        for r in range(start_r + 1, end_r):
            for c in range(start_e, cow_desc + 1):
                brand = ws.cell(r, cow_brand).value.upper().replace("（", "(").replace("）", ")").strip()
                model = ws.cell(r, cow_PN).value.upper().replace("（", "(").replace("）", ")").strip()
                result = df[(df["品牌"] == brand) & (df["型号"] == model)][["品名", "商品描述"]]
                if result.empty:
                    dict_value = {}
                    dict_value["品牌"] = brand
                    dict_value["型号"] = model
                    dict_value["货物中文名称"] = ws.cell(r, cow_brand + 1).value
                    dict_value["单位"] = "个"
                    dict_value["产地"] = ws.cell(r, cow_PN + 1).value
                    dict_value["商品描述"] = ""
                    list_obj.append(dict_value)
                else:
                    ws.cell(r, cow_shopsmall).value = result.iloc[0]["品名"]
                    ws.cell(r, cow_shopname).value = result.iloc[0]["品名"]
                    ws.cell(r, cow_desc).value = result.iloc[0]["商品描述"]

        # print(list_obj)

        time_strappend = self.time_strappend(path)

        if list_obj:
            if len(list_obj) > 1:
                for r in list_obj:
                    isExists = False
                    if self.__set_list == []:
                        isExists = True
                        self.__set_list.append(r)
                    else:
                        for c in self.__set_list:
                            if r["品牌"] == c["品牌"] and r["型号"] == c["型号"]:
                                isExists = True
                                # self.__set_list.append(r)
                    if not isExists:
                        self.__set_list.append(r)
            else:
                self.__set_list = list_obj

            f = open(r"D:\config\filepath.txt", "a", encoding="utf-8")
            f.write(time_strappend + "\n")
            f.close()
            # wb.save(path)
            return self.add_data(time_strappend, wb)

        else:
            wb.save(time_strappend)
            wb.close()
            return time_strappend

    id = 1

    def add_data(self, new_path, wb):
        for item in self.__set_list:
            newmodel = NewModel()
            newmodel.shopname = item["货物中文名称"]
            newmodel.brand = item["品牌"]
            newmodel.model = item["型号"]
            newmodel.place = item["产地"]
            newmodel.unit = item["单位"]
            newmodel.shopdescription = item["商品描述"]
            newmodel.uid = AirModel.id
            self.__newmodel_list.append(newmodel)
            AirModel.id += 1
        return self.__write_to_excel(self.__newmodel_list, new_path, wb)

    def __write_to_excel(self, __newmodel_list, new_path, wb):
        ins = pd.DataFrame(columns=["序号", "品牌", "货物中文名称", "单位", "型号", "产地","商品描述"])
        for i in range(len(__newmodel_list)):
            data = [__newmodel_list[i].uid, __newmodel_list[i].brand, __newmodel_list[i].shopname,
                    __newmodel_list[i].unit,
                    __newmodel_list[i].model, __newmodel_list[i].place,__newmodel_list[i].shopdescription]
            ins.loc[i + 1] = data
        new_path_to_excel = "D:\\Mail\\sendexcel" + "\\" + new_path.split("\\")[2] + "_" + new_path.split("-")[
            -2] + "_" + \
                            new_path.split("-")[-1].split(".")[0] + "_" + "newtable.xlsx"
        writer = pd.ExcelWriter(new_path_to_excel, engine="xlsxwriter")
        ins.to_excel(writer, sheet_name="Sheet1", index=False)  # 写入excel,新型号excel
        wb.save(new_path)  # 原文件excel
        wb.close()
        # df.to_excel(new_path, index=False)  # 原文件excel
        worksheet = writer.sheets["Sheet1"]
        worksheet.set_column("B:Z", 13)  # 更改单元格宽度
        writer.save()
        return new_path_to_excel

    def time_strftime(self):
        return time.strftime("%Y%m%d-%H%M%S")

    def time_strappend(self, file):
        path_list = file.rsplit(".", 1)
        path_list[0] = path_list[0] + "-" + self.time_strftime()
        return ".".join(path_list)


# path = r"C:\Users\windo\Desktop\JL81021038SZ(1.22)HK(1).xlsx"
# model = AirModel()
# print(model.invoke_openpyxl(path))

